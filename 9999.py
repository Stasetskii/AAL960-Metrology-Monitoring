#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import serial
import serial.tools.list_ports
import struct
import time
import json
import os
from functools import reduce
import threading
from datetime import datetime
import winsound

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import LineChart, Reference


ADDR = 0x01
BAUDRATES = [9600, 19200, 38400, 57600, 115200]
START_FRAME = b"\x55\x55\x01\x03\x20\x23\x00\x01\xAA\xAA"

UNITS_P = {
    0x00: 'мм вод.ст.',
    0x01: 'мм рт.ст.',
    0x02: 'мбар',
    0x03: 'бар',
    0x04: 'psi',
    0x05: 'Па',
    0x06: 'МПа',
    0x07: 'кПа',
    0x0A: 'inHg',
    0x0B: 'inH2O',
    0x0C: 'kg/cm²',
}
UNITS_E = {
    0x08: 'мА',
    0x09: 'В',
}

CONFIG_FILE = "stmp960_config.json"


def cs(data):
    return reduce(lambda x, y: x ^ y, data, 0)


def read_frame(ser):
    while ser.in_waiting >= 20:
        if ser.read(2) != b'\x55\x55':
            continue
        hdr = ser.read(2)
        if len(hdr) < 2:
            return None
        addr, length = hdr
        if addr != ADDR:
            continue
        payload = ser.read(length + 3)
        if len(payload) < length + 3:
            return None
        data, recv_cs, end = payload[:length], payload[length], payload[length + 1:]
        if end != b'\xAA\xAA':
            continue
        if recv_cs != cs([ADDR, length] + list(data)):
            continue
        return data
    return None


class CalibratorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("STMP-960 — Профессиональная поверка и мониторинг")
        self.root.geometry("1700x950")
        self.root.minsize(1400, 750)

        self.style = ttk.Style()
        self.style.configure("Fix.TButton", background="#5cb85c", foreground="black")
        self.style.configure("Reset.TButton", background="#d9534f", foreground="black")
        self.style.configure("Undo.TButton", background="#f0ad4e", foreground="black")

        self.ser = None
        self.running = False
        self.poll_thread = None

        self.current_p = 0.0
        self.current_p_unit = "—"
        self.current_signal = 0.0
        self.current_signal_unit = "—"
        self.current_mode = "—"
        self.current_sw_state = None

        self.points_plan = []
        self.calib_points = []
        self.calib_t0 = None  # время начала поверки

        self.monitor_enabled = False
        self.monitor_paused = False
        self.monitor_t0 = None
        self.monitor_times = []
        self.monitor_pressures = []
        self.monitor_signals = []
        self.monitor_pg = []
        self.monitor_min = None
        self.monitor_max = None
        self.last_plot_update = 0

        self.calib_cursor_line = None
        self.mon_cursor_line = None

        self.calib_ax2 = None  # правая ось графика поверки

        self.last_port = ""
        self.last_baud = 9600
        self.last_geometry = "1700x950"

        self.load_config()
        self._build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

        self.root.bind("<F1>", lambda e: self.connect())
        self.root.bind("<F2>", lambda e: self.fix_point())
        self.root.bind("<Delete>", lambda e: self.undo_last_point())

    def load_config(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                    self.last_port = cfg.get("port", "")
                    self.last_baud = cfg.get("baud", 9600)
                    self.last_geometry = cfg.get("geometry", "1700x950")
                    self.root.geometry(self.last_geometry)
            except:
                pass

    def save_config(self):
        cfg = {"port": self.port_combo.get(), "baud": int(self.baud_combo.get()), "geometry": self.root.geometry()}
        try:
            with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(cfg, f)
        except:
            pass

    def _build_ui(self):
        frm_conn = ttk.LabelFrame(self.root, text=" Подключение ")
        frm_conn.pack(fill="x", padx=10, pady=8)

        ttk.Label(frm_conn, text="Порт:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.port_combo = ttk.Combobox(frm_conn, width=15)
        self.port_combo.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        ttk.Button(frm_conn, text="↺", width=3, command=self.refresh_ports).grid(row=0, column=2, padx=2)

        ttk.Label(frm_conn, text="Бодрейт:").grid(row=0, column=3, padx=5, pady=5, sticky="e")
        self.baud_combo = ttk.Combobox(frm_conn, values=BAUDRATES, width=10, state="readonly")
        self.baud_combo.set(self.last_baud or 9600)
        self.baud_combo.grid(row=0, column=4, padx=5, pady=5)

        self.btn_connect = ttk.Button(frm_conn, text="Подключиться (F1)", command=self.connect)
        self.btn_connect.grid(row=1, column=0, columnspan=2, padx=5, pady=5, sticky="we")

        self.btn_disconnect = ttk.Button(frm_conn, text="Отключиться", command=self.disconnect, state="disabled")
        self.btn_disconnect.grid(row=1, column=2, columnspan=3, padx=5, pady=5, sticky="we")

        self.status_var = tk.StringVar(value="Отключено")
        self.lbl_status = ttk.Label(frm_conn, textvariable=self.status_var, foreground="red", font=("", 10, "bold"))
        self.lbl_status.grid(row=0, column=5, padx=20, pady=5)

        self.refresh_ports()

        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self._build_calibration_tab()
        self._build_monitor_tab()

    def refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.port_combo["values"] = ports
        if ports:
            if self.last_port in ports:
                self.port_combo.set(self.last_port)
            elif not self.port_combo.get():
                self.port_combo.set(ports[0])

    # ================= ТАБ «ПОВЕРКА» =================

    def _build_calibration_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="Поверка")

        # текущие показания
        frm_live = ttk.LabelFrame(tab, text=" Текущие показания ")
        frm_live.pack(fill="x", padx=10, pady=5)

        ttk.Label(frm_live, text="Режим:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
        self.mode_var = tk.StringVar(value="—")
        ttk.Label(frm_live, textvariable=self.mode_var, font=("Segoe UI", 11, "bold")).grid(
            row=0, column=1, padx=5, pady=5, sticky="w"
        )

        ttk.Label(frm_live, text="Давление (эталон):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
        self.p_var = tk.StringVar(value="0.000000")
        self.p_unit_var = tk.StringVar(value="—")
        ttk.Label(frm_live, textvariable=self.p_var, font=("Consolas", 14, "bold")).grid(
            row=1, column=1, padx=5, pady=5, sticky="w"
        )
        ttk.Label(frm_live, textvariable=self.p_unit_var).grid(row=1, column=2, padx=2, sticky="w")

        ttk.Label(frm_live, text="Сигнал от датчика:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
        self.signal_var = tk.StringVar(value="0.000000")
        self.signal_unit_var = tk.StringVar(value="—")
        ttk.Label(frm_live, textvariable=self.signal_var, font=("Consolas", 12)).grid(
            row=2, column=1, padx=5, pady=5, sticky="w"
        )
        ttk.Label(frm_live, textvariable=self.signal_unit_var).grid(row=2, column=2, padx=2, sticky="w")

        self.btn_fix_point = ttk.Button(
            frm_live,
            text="Зафиксировать точку (F2)",
            command=self.fix_point,
            style="Fix.TButton",
            state="disabled",
            width=28,
        )
        self.btn_fix_point.grid(row=0, column=4, padx=10, pady=5, sticky="e")

        self.btn_undo_point = ttk.Button(
            frm_live,
            text="↩ Откатить точку (Del)",
            command=self.undo_last_point,
            style="Undo.TButton",
            state="disabled",
            width=28,
        )
        self.btn_undo_point.grid(row=1, column=4, padx=10, pady=5, sticky="e")

        # поверяемый датчик
        frm_dut = ttk.LabelFrame(tab, text=" Поверяемый датчик ")
        frm_dut.pack(fill="x", padx=10, pady=5)

        self.dut_inner_id = tk.StringVar()
        self.dut_serial = tk.StringVar()
        self.dut_model = tk.StringVar()
        self.dut_range_low = tk.StringVar(value="0")
        self.dut_range_high = tk.StringVar(value="10")
        self.dut_max_error = tk.StringVar(value="0.5")
        self.dut_points_count = tk.StringVar(value="5")
        self.dut_reverse = tk.BooleanVar(value=True)

        row = 0
        ttk.Label(frm_dut, text="Внутр. №:").grid(row=row, column=0, padx=5, pady=3, sticky="e")
        ttk.Entry(frm_dut, textvariable=self.dut_inner_id, width=15).grid(row=row, column=1, padx=5, pady=3, sticky="w")
        ttk.Label(frm_dut, text="Зав. №:").grid(row=row, column=2, padx=5, pady=3, sticky="e")
        ttk.Entry(frm_dut, textvariable=self.dut_serial, width=20).grid(row=row, column=3, padx=5, pady=3, sticky="w")

        row += 1
        ttk.Label(frm_dut, text="Модель:").grid(row=row, column=0, padx=5, pady=3, sticky="e")
        ttk.Entry(frm_dut, textvariable=self.dut_model, width=20).grid(row=row, column=1, padx=5, pady=3, sticky="w")
        ttk.Label(frm_dut, text="Диапазон:").grid(row=row, column=2, padx=5, pady=3, sticky="e")
        ttk.Entry(frm_dut, textvariable=self.dut_range_low, width=8).grid(row=row, column=3, padx=5, pady=3, sticky="w")
        ttk.Label(frm_dut, text="…").grid(row=row, column=4, padx=2)
        ttk.Entry(frm_dut, textvariable=self.dut_range_high, width=8).grid(row=row, column=5, padx=5, pady=3, sticky="w")
        ttk.Label(frm_dut, text="±ПГ, %:").grid(row=row, column=6, padx=5, pady=3, sticky="e")
        ttk.Entry(frm_dut, textvariable=self.dut_max_error, width=6).grid(row=row, column=7, padx=5, pady=3, sticky="w")

        row += 1
        ttk.Label(frm_dut, text="Точек вверх:").grid(row=row, column=0, padx=5, pady=3, sticky="e")
        ttk.Combobox(frm_dut, textvariable=self.dut_points_count, values=[str(i) for i in range(3, 12)],
                     width=5, state="readonly").grid(row=row, column=1, padx=5, pady=3, sticky="w")
        ttk.Checkbutton(frm_dut, text="Обратный ход", variable=self.dut_reverse).grid(
            row=row, column=2, columnspan=4, padx=5, pady=3, sticky="w"
        )

        # основной сплиттер: таблица + кнопки + график
        frm_main = ttk.PanedWindow(tab, orient="horizontal")
        frm_main.pack(fill="both", expand=True, padx=10, pady=5)

        # таблица
        frm_table = ttk.LabelFrame(frm_main, text=" Протокол поверки ")
        frm_main.add(frm_table, weight=2)

        columns = ("№", "Время", "Рточки", "Рэт", "Сигнал", "Ррасч", "ПГ (%)", "Соответствие")
        self.tree = ttk.Treeview(frm_table, columns=columns, show="headings", height=20)
        widths = [50, 90, 110, 110, 100, 110, 100, 140]
        for i, (text, w) in enumerate(zip(columns, widths)):
            self.tree.heading(i, text=text)
            self.tree.column(i, width=w, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(frm_table, orient="vertical", command=self.tree.yview)
        sb.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=sb.set)

        self.tree.tag_configure("ok", background="#d4edda")
        self.tree.tag_configure("fail", background="#f8d7da")
        self.tree.tag_configure("future", foreground="gray")

        # кнопки справа
        frm_btns = ttk.Frame(frm_main)
        frm_main.add(frm_btns, weight=0)
        ttk.Button(frm_btns, text="Экспорт в Excel", command=self.export_calib_to_excel).pack(pady=5)
        ttk.Button(frm_btns, text="Очистить протокол", command=self.clear_points, style="Reset.TButton").pack(pady=5)

        # график поверки
        frm_plot = ttk.LabelFrame(frm_main, text=" График поверки (Рточки / Ррасч / Рэт) ")
        frm_main.add(frm_plot, weight=3)

        self.calib_fig = Figure(figsize=(8, 6), dpi=100)
        self.calib_ax = self.calib_fig.add_subplot(111)
        self.calib_ax2 = self.calib_ax.twinx()  # правая ось создаётся ОДИН раз

        self.calib_ax.grid(True)
        self.calib_ax.set_xlabel("Время, с")
        self.calib_ax.set_ylabel("Рточки / Рэт")
        self.calib_ax2.set_ylabel("Ррасч")

        self.calib_canvas = FigureCanvasTkAgg(self.calib_fig, frm_plot)
        toolbar = NavigationToolbar2Tk(self.calib_canvas, frm_plot)
        toolbar.update()
        self.calib_canvas.get_tk_widget().pack(fill="both", expand=True)

        self.calib_cursor_var = tk.StringVar(value="Курсор: —")
        ttk.Label(frm_plot, textvariable=self.calib_cursor_var).pack(side="bottom", anchor="w", padx=5)

        self.calib_canvas.mpl_connect("motion_notify_event", self.on_calib_motion)
        self.calib_canvas.mpl_connect("button_press_event", self.on_calib_click)

        for var in (self.dut_range_low, self.dut_range_high, self.dut_points_count, self.dut_reverse):
            var.trace_add("write", self.update_plan)

        self.update_plan()
        self.update_table_headers()

    # ================= ТАБ «МОНИТОРИНГ» =================

    def _build_monitor_tab(self):
        tab = ttk.Frame(self.nb)
        self.nb.add(tab, text="Мониторинг")

        top = ttk.Frame(tab)
        top.pack(fill="x", padx=5, pady=5)

        frm_info = ttk.LabelFrame(top, text=" Статистика ")
        frm_info.pack(side="left", fill="x", expand=True)

        self.mon_p_var = tk.StringVar(value="0.000000")
        self.mon_min_var = tk.StringVar(value="—")
        self.mon_max_var = tk.StringVar(value="—")
        self.mon_span_var = tk.StringVar(value="—")
        self.mon_sw_state_var = tk.StringVar(value="NA")
        self.mon_p_unit_var = tk.StringVar(value="—")

        labels = ["Текущее:", "Минимум:", "Максимум:", "ΔP:", "Контакт:"]
        vars_list = [self.mon_p_var, self.mon_min_var, self.mon_max_var, self.mon_span_var, self.mon_sw_state_var]

        for i, (text, var) in enumerate(zip(labels, vars_list)):
            ttk.Label(frm_info, text=text).grid(row=i, column=0, padx=5, pady=4, sticky="e")
            if i == 0:
                ttk.Label(frm_info, textvariable=var, font=("Consolas", 12)).grid(row=i, column=1, padx=5, sticky="w")
                ttk.Label(frm_info, textvariable=self.mon_p_unit_var).grid(row=i, column=2, padx=2)
            else:
                ttk.Label(frm_info, textvariable=var).grid(row=i, column=1, columnspan=2, padx=5, sticky="w")

        frm_btn = ttk.Frame(top)
        frm_btn.pack(side="right", padx=20)

        self.btn_mon_start = ttk.Button(frm_btn, text="Старт (Space)", command=self.start_monitoring, state="disabled")
        self.btn_mon_start.pack(fill="x", pady=2)

        self.btn_pause = ttk.Button(frm_btn, text="Пауза", command=self.toggle_monitor_pause, state="disabled")
        self.btn_pause.pack(fill="x", pady=2)

        self.btn_mon_reset = ttk.Button(frm_btn, text="Сброс", command=self.reset_monitoring, state="disabled")
        self.btn_mon_reset.pack(fill="x", pady=2)

        self.btn_mon_export = ttk.Button(
            frm_btn, text="Экспорт в Excel", command=self.export_monitor_to_excel, state="disabled"
        )
        self.btn_mon_export.pack(fill="x", pady=2)

        frm_plot = ttk.LabelFrame(tab, text=" Мониторинг по времени ")
        frm_plot.pack(fill="both", expand=True, padx=5, pady=5)

        self.mon_fig = Figure(figsize=(8, 4), dpi=100)
        self.mon_ax = self.mon_fig.add_subplot(111)
        self.mon_ax.grid(True)

        self.mon_canvas = FigureCanvasTkAgg(self.mon_fig, frm_plot)
        toolbar = NavigationToolbar2Tk(self.mon_canvas, frm_plot)
        toolbar.update()
        self.mon_canvas.get_tk_widget().pack(fill="both", expand=True)

        self.mon_cursor_var = tk.StringVar(value="Курсор: —")
        ttk.Label(frm_plot, textvariable=self.mon_cursor_var).pack(side="bottom", anchor="w", padx=5)

        self.mon_canvas.mpl_connect("motion_notify_event", self.on_mon_motion)
        self.mon_canvas.mpl_connect("button_press_event", self.on_mon_click)

    # =============== ПЛАН ТОЧЕК, ШАПКА ТАБЛИЦЫ ===============

    def build_target_plan(self):
        try:
            low = float(self.dut_range_low.get())
            high = float(self.dut_range_high.get())
            n = int(self.dut_points_count.get())
            if n < 2 or high == low:
                return []
            step = (high - low) / (n - 1)
            forward = [round(low + i * step, 6) for i in range(n)]
            if self.dut_reverse.get():
                reverse = forward[-2::-1]
                return forward + reverse
            return forward
        except:
            return []

    def update_plan(self, *args):
        self.points_plan = self.build_target_plan()
        self.calib_points.clear()
        self.calib_t0 = None  # сброс времени начала поверки
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, p_set in enumerate(self.points_plan):
            self.tree.insert("", "end", values=(i+1, "—", f"{p_set:.6f}", "—", "—", "—", "—", "—"), tags=("future",))
        self.update_calib_plot()

    def update_table_headers(self):
        p_unit = self.current_p_unit if self.current_p_unit not in ("—", None, "") else ""
        s_unit = self.current_signal_unit if self.current_signal_unit not in ("—", None, "") else ""

        def col_title(base, unit):
            return f"{base} / {unit}" if unit else base

        self.tree.heading(0, text="№")
        self.tree.heading(1, text="Время")
        self.tree.heading(2, text=col_title("Рточки", p_unit))
        self.tree.heading(3, text=col_title("Рэт", p_unit))
        self.tree.heading(4, text=col_title("Сигнал", s_unit))
        self.tree.heading(5, text=col_title("Ррасч", p_unit))
        self.tree.heading(6, text="ПГ, %")
        self.tree.heading(7, text="Соответствие")

    # =============== РАБОТА С ТОЧКАМИ ===============

    def fix_point(self):
        if len(self.calib_points) >= len(self.points_plan):
            return

        idx = len(self.calib_points)
        p_set = self.points_plan[idx] if idx < len(self.points_plan) else self.current_p

        low = float(self.dut_range_low.get() or 0)
        high = float(self.dut_range_high.get() or 0)
        span = abs(high - low) or 1.0

        if self.current_mode == "I/P":
            p_calc = low + (self.current_signal - 4.0) * span / 16.0
        elif self.current_mode == "V/P":
            p_calc = low + self.current_signal * span / 10.0
        else:
            p_calc = None

        pg = abs(p_calc - self.current_p) / span * 100 if p_calc is not None else None
        max_err = float(self.dut_max_error.get() or 0)
        ok = "Соответствует" if pg is None or pg <= max_err else "Не соответствует"

        if self.calib_t0 is None:
            self.calib_t0 = time.time()
        t_rel = time.time() - self.calib_t0

        point = {
            "idx": idx + 1,
            "time": datetime.now().strftime("%H:%M:%S"),
            "p_set": p_set,
            "p_et": self.current_p,
            "signal": self.current_signal,
            "p_calc": p_calc,
            "pg": pg,
            "ok": ok,
            "t": t_rel,
        }
        self.calib_points.append(point)

        item = self.tree.get_children()[idx]
        self.tree.item(
            item,
            values=(
                point["idx"],
                point["time"],
                f"{p_set:.6f}",
                f"{self.current_p:.6f}",
                f"{self.current_signal:.6f}",
                f"{p_calc:.6f}" if p_calc is not None else "—",
                f"{pg:.3f}" if pg is not None else "—",
                ok,
            ),
            tags=("ok" if ok == "Соответствует" else "fail",),
        )

        try:
            winsound.Beep(800, 150)
        except:
            pass

        self.update_calib_plot()

        if len(self.calib_points) == len(self.points_plan):
            verdict = "Соответствует" if all(p["ok"] == "Соответствует" for p in self.calib_points) else "Не соответствует"
            messagebox.showinfo("Поверка завершена", f"Все точки зафиксированы!\n\nЗаключение: {verdict}")

    def undo_last_point(self):
        if not self.calib_points:
            return
        self.calib_points.pop()
        if not self.calib_points:
            self.calib_t0 = None
        self.refresh_table_from_calib_points()
        self.update_calib_plot()

    def refresh_table_from_calib_points(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for i, p_set in enumerate(self.points_plan):
            if i < len(self.calib_points):
                pt = self.calib_points[i]
                ok_tag = "ok" if pt["ok"] == "Соответствует" else "fail"
                self.tree.insert("", "end", values=(
                    pt["idx"], pt["time"],
                    f"{p_set:.6f}", f"{pt['p_et']:.6f}",
                    f"{pt['signal']:.6f}",
                    f"{pt['p_calc']:.6f}" if pt['p_calc'] is not None else "—",
                    f"{pt['pg']:.3f}" if pt['pg'] is not None else "—",
                    pt["ok"]
                ), tags=(ok_tag,))
            else:
                self.tree.insert("", "end", values=(i+1, "—", f"{p_set:.6f}", "—", "—", "—", "—", "—"), tags=("future",))

    def clear_points(self):
        if messagebox.askyesno("Очистка", "Очистить протокол поверки?"):
            self.calib_points.clear()
            self.calib_t0 = None
            self.update_plan()

    def _get_point_time(self, idx: int) -> float:
        """Время точки для оси X (сек от начала поверки)."""
        pt = self.calib_points[idx]
        t = pt.get("t")
        if t is None:
            return float(idx)
        return float(t)

    def update_calib_plot(self):
        # X — время, слева — Рточки/Рэт, справа — Ррасч
        # ЧИСТИМ обе оси, не создаём новые
        self.calib_ax.cla()
        self.calib_ax2.cla()

        self.calib_ax.set_xlabel("Время, с")
        self.calib_ax.set_ylabel("Рточки / Рэт")
        self.calib_ax.grid(True)
        self.calib_ax2.set_ylabel("Ррасч")

        if self.calib_points:
            t_vals = [self._get_point_time(i) for i in range(len(self.calib_points))]
            p_set_vals = [p["p_set"] for p in self.calib_points]
            p_et_vals = [p["p_et"] for p in self.calib_points]

            calc_pairs = [
                (self._get_point_time(i), p["p_calc"])
                for i, p in enumerate(self.calib_points)
                if p["p_calc"] is not None
            ]

            if t_vals:
                self.calib_ax.plot(t_vals, p_set_vals, "o-", linewidth=2, markersize=6, label="Рточки")
                self.calib_ax.plot(t_vals, p_et_vals, "x--", linewidth=1.5, markersize=6, label="Рэт")

            if calc_pairs:
                t_calc, p_calc_vals = zip(*calc_pairs)
                self.calib_ax2.plot(t_calc, p_calc_vals, "s-.", linewidth=1.5, markersize=5, label="Ррасч")

            lines1, labels1 = self.calib_ax.get_legend_handles_labels()
            lines2, labels2 = self.calib_ax2.get_legend_handles_labels()
            if lines1 or lines2:
                self.calib_ax.legend(lines1 + lines2, labels1 + labels2, loc="best")

        self.calib_fig.tight_layout()
        self.calib_canvas.draw_idle()

    def on_calib_motion(self, event):
        if not event.inaxes or not self.calib_points:
            self.calib_cursor_var.set("Курсор: —")
            return
        x = event.xdata
        idx = min(range(len(self.calib_points)), key=lambda i: abs(self._get_point_time(i) - x))
        pt = self.calib_points[idx]
        t = self._get_point_time(idx)
        p_calc_str = f"{pt['p_calc']:.6f}" if pt['p_calc'] is not None else "—"
        pg_str = f"{pt['pg']:.3f}" if pt['pg'] is not None else "—"
        self.calib_cursor_var.set(
            f"t={t:.1f} с | Точка {pt['idx']} | Рточки={pt['p_set']:.6f} | Рэт={pt['p_et']:.6f} | "
            f"Сигнал={pt['signal']:.6f} | Ррасч={p_calc_str} | ПГ={pg_str}% | {pt['ok']}"
        )
        if self.calib_cursor_line is None:
            self.calib_cursor_line = self.calib_ax.axvline(t, color="gray", linestyle="--", alpha=0.7)
        else:
            self.calib_cursor_line.set_xdata([t])
        self.calib_canvas.draw_idle()

    def on_calib_click(self, event):
        if event.inaxes and self.calib_points:
            x = event.xdata
            idx = min(range(len(self.calib_points)), key=lambda i: abs(self._get_point_time(i) - x))
            t = self._get_point_time(idx)
            if self.calib_cursor_line:
                self.calib_cursor_line.set_color("red")
                self.calib_cursor_line.set_linewidth(2)
                self.calib_cursor_line.set_xdata([t])
            else:
                self.calib_cursor_line = self.calib_ax.axvline(t, color="red", linestyle="--", linewidth=2)
            self.calib_canvas.draw_idle()

    def export_calib_to_excel(self):
        if not self.calib_points:
            messagebox.showwarning("Нет данных", "Зафиксируйте точки")
            return

        filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
            initialfile=f"протокол_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not filename:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Протокол"

        ws.append(["ПРОТОКОЛ ПОВЕРКИ ДАТЧИКА ДАВЛЕНИЯ"])
        ws.merge_cells('A1:H1')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal="center")

        ws.append([f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"])
        ws.append([f"Датчик: {self.dut_model.get()} | Зав. № {self.dut_serial.get()} | Внутр. № {self.dut_inner_id.get()}"])
        ws.append([f"Диапазон: {self.dut_range_low.get()} … {self.dut_range_high.get()} {self.current_p_unit}"])
        ws.append([f"Сигнал: {self.current_signal_unit}"])
        ws.append([f"Допустимая приведённая погрешность: ±{self.dut_max_error.get()} %"])
        ws.append([])

        p_unit = self.current_p_unit if self.current_p_unit not in ("—", None, "") else ""
        s_unit = self.current_signal_unit if self.current_signal_unit not in ("—", None, "") else ""

        def h_title(base, unit):
            return f"{base} / {unit}" if unit else base

        headers = [
            "№",
            "Время",
            h_title("Рточки", p_unit),
            h_title("Рэт", p_unit),
            h_title("Сигнал", s_unit),
            h_title("Ррасч", p_unit),
            "ПГ, %",
            "Соответствие",
        ]
        ws.append(headers)
        for cell in ws[ws.max_row]:
            if not isinstance(cell, MergedCell):
                cell.font = Font(bold=True)

        for pt in self.calib_points:
            row = [
                pt["idx"], pt["time"],
                f"{pt['p_set']:.6f}", f"{pt['p_et']:.6f}",
                f"{pt['signal']:.6f}", f"{pt['p_calc']:.6f}" if pt['p_calc'] is not None else "—",
                f"{pt['pg']:.3f}" if pt['pg'] is not None else "—", pt["ok"]
            ]
            ws.append(row)
            if pt["ok"] == "Не соответствует":
                for cell in ws[ws.max_row]:
                    if not isinstance(cell, MergedCell):
                        cell.fill = PatternFill("solid", fgColor="ffcccc")

        verdict = "Соответствует" if all(p["ok"] == "Соответствует" for p in self.calib_points) else "Не соответствует"
        ws.append([])
        ws.append(["ЗАКЛЮЧЕНИЕ:", verdict])
        ws[f"B{ws.max_row}"].font = Font(size=14, bold=True, color="006400" if verdict == "Соответствует" else "FF0000")

        for i, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                if cell.value and not isinstance(cell, MergedCell):
                    max_len = max(max_len, len(str(cell.value)))
            column = get_column_letter(i)
            ws.column_dimensions[column].width = min(max_len + 2, 50)

        wb.save(filename)
        messagebox.showinfo("Готово", f"Протокол сохранён\n{filename}\n\nЗаключение: {verdict}")

    def export_monitor_to_excel(self):
        if not self.monitor_times:
            messagebox.showwarning("Нет данных", "Запустите мониторинг")
            return
        filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
            initialfile=f"мониторинг_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not filename:
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Данные"

        ws.append(["#", "Время, с", "Давление (эталон)", "Сигнал", "ПГ (%)"])
        for i, (t, p, s, pg) in enumerate(zip(self.monitor_times, self.monitor_pressures, self.monitor_signals, self.monitor_pg), 1):
            ws.append([i, round(t, 3), round(p, 6), round(s, 6), round(pg, 3)])

        chart_sheet = wb.create_sheet("График")
        chart = LineChart()
        chart.title = "Мониторинг"
        chart.x_axis.title = "Время, с"
        chart.y_axis.title = "Давление (эталон)"

        data = Reference(ws, min_col=3, max_col=3, min_row=1, max_row=len(self.monitor_times) + 1)
        cats = Reference(ws, min_col=2, min_row=2, max_row=len(self.monitor_times) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart_sheet.add_chart(chart, "A1")

        wb.save(filename)
        messagebox.showinfo("Готово", f"Данные мониторинга сохранены\n{filename}")

    # =============== ПОДКЛЮЧЕНИЕ И ОПРОС ===============

    def connect(self):
        if self.ser:
            return
        port = self.port_combo.get()
        try:
            baud = int(self.baud_combo.get())
        except:
            messagebox.showerror("Ошибка", "Выберите скорость")
            return

        try:
            tmp = serial.Serial(port, baud, timeout=0.1)
            time.sleep(1.5)
            tmp.write(START_FRAME)
            tmp.close()
        except Exception as e:
            messagebox.showwarning("Пинок", f"Не удалось отправить старт:\n{e}")

        time.sleep(0.5)

        try:
            self.ser = serial.Serial(port, baud, timeout=0.2)
        except Exception as e:
            messagebox.showerror("Порт", str(e))
            return

        self.running = True
        self.poll_thread = threading.Thread(target=self.poll_loop, daemon=True)
        self.poll_thread.start()

        self.btn_connect.config(state="disabled")
        self.btn_disconnect.config(state="normal")
        self.btn_fix_point.config(state="normal")
        self.btn_undo_point.config(state="normal")

        self.btn_mon_start.config(state="normal")
        self.btn_mon_reset.config(state="normal")
        self.btn_mon_export.config(state="normal")

        self.status_var.set("Подключено")
        self.lbl_status.config(foreground="green")
        self.save_config()

    def disconnect(self):
        self.running = False
        if self.ser:
            try:
                self.ser.close()
            except:
                pass
            self.ser = None

        self.btn_connect.config(state="normal")
        self.btn_disconnect.config(state="disabled")
        self.btn_fix_point.config(state="disabled")
        self.btn_undo_point.config(state="disabled")

        self.btn_mon_start.config(state="disabled")
        self.btn_pause.config(state="disabled", text="Пауза")
        self.btn_mon_reset.config(state="disabled")
        self.btn_mon_export.config(state="disabled")

        self.status_var.set("Отключено")
        self.lbl_status.config(foreground="red")
        self.reset_monitoring()
        self.clear_points()

    def poll_loop(self):
        while self.running:
            try:
                data = read_frame(self.ser)
                if data and len(data) == 13:
                    header = data[0:3]
                    if header in (b"\x30\x15\x01", b"\x30\x16\x01"):
                        p = struct.unpack(">f", data[3:7])[0]
                        sec = struct.unpack(">f", data[8:12])[0]
                        p_unit = UNITS_P.get(data[7], f"0x{data[7]:02X}")
                        sec_unit = UNITS_E.get(data[12], f"0x{data[12]:02X}")
                        mode = "I/P" if header == b"\x30\x15\x01" else "V/P"
                    elif header == b"\x30\x17\x01":
                        p = struct.unpack(">f", data[3:7])[0]
                        p_unit = UNITS_P.get(data[7], f"0x{data[7]:02X}")
                        code = data[12]
                        if code == 0x03:
                            sec, sec_unit = True, "контакт замкнут"
                        elif code == 0x04:
                            sec, sec_unit = False, "контакт разомкнут"
                        else:
                            sec, sec_unit = code, f"реле 0x{code:02X}"
                        mode = "Реле"
                    else:
                        time.sleep(0.005)
                        continue

                    self.root.after(0, self.update_ui, mode, p, p_unit, sec, sec_unit)
                else:
                    time.sleep(0.005)
            except serial.SerialException:
                self.root.after(0, lambda: messagebox.showerror("Связь", "Порт пропал!"))
                self.root.after(0, self.disconnect)
                break
            except Exception as e:
                print("poll error:", e)
                time.sleep(0.1)

    # =============== ОБНОВЛЕНИЕ UI И МОНИТОРИНГА ===============

    def update_ui(self, mode, p, p_unit, sec, sec_unit):
        self.current_p = p
        self.current_p_unit = p_unit
        self.current_signal = sec if not isinstance(sec, bool) else 0.0
        self.current_signal_unit = sec_unit
        self.current_mode = mode
        self.current_sw_state = sec if isinstance(sec, bool) else None

        self.mode_var.set(mode)
        self.p_var.set(f"{p:.6f}")
        self.p_unit_var.set(p_unit)
        self.signal_var.set(f"{sec:.6f}" if not isinstance(sec, bool) else sec_unit)
        self.signal_unit_var.set(sec_unit if not isinstance(sec, bool) else "")

        self.mon_p_var.set(f"{p:.6f}")
        self.mon_p_unit_var.set(p_unit)
        if self.current_sw_state is True:
            self.mon_sw_state_var.set("замкнут")
        elif self.current_sw_state is False:
            self.mon_sw_state_var.set("разомкнут")
        else:
            self.mon_sw_state_var.set("NA")

        self.update_table_headers()
        self.update_monitor(p)

    def update_monitor(self, p):
        if not self.monitor_enabled or self.monitor_paused:
            return

        if self.monitor_t0 is None:
            self.monitor_t0 = time.time()

        t = time.time() - self.monitor_t0
        self.monitor_times.append(t)
        self.monitor_pressures.append(p)
        self.monitor_signals.append(self.current_signal if not isinstance(self.current_signal, bool) else 0.0)
        self.monitor_pg.append(0.0)

        if len(self.monitor_times) > 5000:
            self.monitor_times = self.monitor_times[-5000:]
            self.monitor_pressures = self.monitor_pressures[-5000:]
            self.monitor_signals = self.monitor_signals[-5000:]
            self.monitor_pg = self.monitor_pg[-5000:]

        if self.monitor_min is None or p < self.monitor_min:
            self.monitor_min = p
        if self.monitor_max is None or p > self.monitor_max:
            self.monitor_max = p

        if self.monitor_min is not None:
            self.mon_min_var.set(f"{self.monitor_min:.6f}")
        else:
            self.mon_min_var.set("—")
        if self.monitor_max is not None:
            self.mon_max_var.set(f"{self.monitor_max:.6f}")
        else:
            self.mon_max_var.set("—")
        if self.monitor_min is not None and self.monitor_max is not None:
            self.mon_span_var.set(f"{self.monitor_max - self.monitor_min:.6f}")
        else:
            self.mon_span_var.set("—")

        now = time.time()
        if now - self.last_plot_update > 0.1:
            self.last_plot_update = now
            self.root.after_idle(self.update_monitor_plot)

    def update_monitor_plot(self):
        self.mon_ax.clear()
        self.mon_ax.set_xlabel("Время, с")

        p_unit = self.current_p_unit if self.current_p_unit not in ("—", None, "") else ""
        ylabel = f"Рэт / {p_unit}" if p_unit else "Рэт"
        self.mon_ax.set_ylabel(ylabel)
        self.mon_ax.grid(True)

        if self.monitor_times:
            self.mon_ax.plot(self.monitor_times, self.monitor_pressures)

            if self.monitor_min is not None and self.monitor_max is not None:
                if self.monitor_min == self.monitor_max:
                    delta = abs(self.monitor_min) * 0.05 or 1.0
                    ymin = self.monitor_min - delta
                    ymax = self.monitor_max + delta
                else:
                    span = self.monitor_max - self.monitor_min
                    margin = span * 0.05
                    ymin = self.monitor_min - margin
                    ymax = self.monitor_max + margin
                self.mon_ax.set_ylim(ymin, ymax)

        self.mon_canvas.draw_idle()

    def on_mon_motion(self, event):
        if not event.inaxes or not self.monitor_times:
            self.mon_cursor_var.set("Курсор: —")
            return

        t = event.xdata
        idx = min(range(len(self.monitor_times)), key=lambda i: abs(self.monitor_times[i] - t))
        t_point = self.monitor_times[idx]
        p = self.monitor_pressures[idx]

        self.mon_cursor_var.set(f"t = {t_point:.1f} с | Рэт = {p:.6f}")

        if self.mon_cursor_line is None:
            self.mon_cursor_line = self.mon_ax.axvline(t_point, linestyle="--", alpha=0.7)
        else:
            self.mon_cursor_line.set_xdata([t_point])
        self.mon_canvas.draw_idle()

    def on_mon_click(self, event):
        if event.inaxes and self.monitor_times and self.mon_cursor_line is not None:
            t = event.xdata
            idx = min(range(len(self.monitor_times)), key=lambda i: abs(self.monitor_times[i] - t))
            t_point = self.monitor_times[idx]
            self.mon_cursor_line.set_xdata([t_point])
            self.mon_cursor_line.set_color("red")
            self.mon_cursor_line.set_linewidth(2)
            self.mon_canvas.draw_idle()

    # =============== УПРАВЛЕНИЕ МОНИТОРИНГОМ ===============

    def start_monitoring(self):
        self.monitor_enabled = True
        self.monitor_paused = False
        self.monitor_t0 = time.time()
        self.monitor_times.clear()
        self.monitor_pressures.clear()
        self.monitor_signals.clear()
        self.monitor_pg.clear()
        self.monitor_min = None
        self.monitor_max = None
        self.mon_min_var.set("—")
        self.mon_max_var.set("—")
        self.mon_span_var.set("—")
        self.btn_pause.config(state="normal", text="Пауза")
        self.update_monitor_plot()

    def toggle_monitor_pause(self):
        self.monitor_paused = not self.monitor_paused
        self.btn_pause.config(text="Продолжить" if self.monitor_paused else "Пауза")

    def reset_monitoring(self):
        self.monitor_enabled = False
        self.monitor_paused = False
        self.monitor_t0 = None
        self.monitor_times.clear()
        self.monitor_pressures.clear()
        self.monitor_signals.clear()
        self.monitor_pg.clear()
        self.monitor_min = None
        self.monitor_max = None
        self.mon_ax.clear()
        self.mon_ax.grid(True)
        self.mon_canvas.draw_idle()
        self.mon_min_var.set("—")
        self.mon_max_var.set("—")
        self.mon_span_var.set("—")
        self.mon_sw_state_var.set("NA")

    # =============== ЗАКРЫТИЕ ОКНА ===============

    def on_close(self):
        self.disconnect()
        self.save_config()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = CalibratorGUI(root)
    root.mainloop()
